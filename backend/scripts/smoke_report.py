"""End-to-end report smoke test.

    PYTHONPATH=. uv run python scripts/smoke_report.py

What it does:
  1. Creates a dummy xlsx template (no template files exist in data/ yet).
  2. Picks two demo images + the Finnish demo audio.
  3. Calls transcribe() on the audio.
  4. Calls analyse_image() on each image (in parallel).
  5. Calls build_report() to synthesise + write the .xlsx.
  6. Reads the output back and prints the sheets/rows.

Cost: 1 Whisper call + 2 vision calls + 1 synthesis call.
"""

from __future__ import annotations

import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from pathlib import Path

from app.domain.inspection import ImageAsset
from app.services.report_builder import build_report
from app.services.transcription import transcribe
from app.services.vision import analyse_image
from openpyxl import Workbook, load_workbook

AUDIO_FILE = Path(
    "../data/2 Inspection Audios/R&I Kyllikki Hall Inspection Demo Audio/"
    "R&I Kyllikki Hall Inspection Finnish.m4a"
)
IMAGE_DIR = Path("../data/1 Inspection Pictures/R&I Kyllikki Hall Inspection Demo Pictures")


def make_dummy_template(path: Path) -> Path:
    """Write a minimal xlsx template the LLM can mirror."""
    wb = Workbook()
    ws_inspection = wb.active
    if ws_inspection is None:
        raise RuntimeError("openpyxl returned no active sheet")
    ws_inspection.title = "Inspection"
    ws_inspection.append(
        [
            "Image",
            "Captured At",
            "Location",
            "Component",
            "Condition",
            "Severity",
            "Observation",
            "Recommendation",
        ]
    )
    # Sample row to give the LLM a shape to copy.
    ws_inspection.append(
        [
            "DJI_example.JPG",
            "2026-01-01T00:00:00",
            "Hall A",
            "Pillar weld",
            "Surface paint flaking",
            "minor",
            "Localised paint loss on lower bracket; no structural damage visible.",
            "Repaint within 6 months.",
        ]
    )

    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Inspection Summary"])
    ws_summary.append(["Site", ""])
    ws_summary.append(["Date", ""])
    ws_summary.append(["Inspector", ""])
    ws_summary.append(["Overall severity", ""])
    ws_summary.append(["Notes", ""])

    wb.save(path)
    return path


def parse_dji_capture_time(filename: str) -> datetime:
    """Parse DJI_YYYYMMDDHHMMSS_xxxx_V.JPG → datetime."""
    stem = Path(filename).stem  # DJI_20260317154903_0015_V
    ts = stem.split("_")[1]  # 20260317154903
    return datetime.strptime(ts, "%Y%m%d%H%M%S")


def main() -> int:
    if not AUDIO_FILE.exists():
        print(f"FAIL: audio missing: {AUDIO_FILE.resolve()}")
        return 1
    images = sorted(IMAGE_DIR.glob("*.JPG"))[:2]
    if len(images) < 2:
        print(f"FAIL: need at least 2 images in {IMAGE_DIR}")
        return 1

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)
        template_path = make_dummy_template(tmp_dir / "dummy_template.xlsx")
        output_path = tmp_dir / "report.xlsx"

        print(f"Template: {template_path}")
        print(f"Images  : {[p.name for p in images]}")
        print(f"Audio   : {AUDIO_FILE.name}")

        # --- Build ImageAssets with metadata derived from the DJI filenames ---
        image_assets = [
            ImageAsset(
                filename=p.name,
                path=p,
                captured_at=parse_dji_capture_time(p.name),
                location=f"Hall A · Pillar {10 + i}",
            )
            for i, p in enumerate(images)
        ]

        # --- 1. Whisper ---
        print("\n[1/3] Whisper transcribe (audio_language=fi)…")
        transcript = transcribe(AUDIO_FILE, language="fi")
        print(f"      lang={transcript.language} segs={len(transcript.segments)}")

        # --- 2. Vision (parallel) ---
        print("\n[2/3] Vision analyse_image (x2 parallel)...")
        with ThreadPoolExecutor(max_workers=2) as pool:
            findings = list(
                pool.map(
                    lambda img: analyse_image(
                        img.path,
                        operator_commentary=transcript.text[:600],
                        location=img.location,
                    ),
                    image_assets,
                )
            )
        for f in findings:
            print(f"      {f.image}: {f.severity} · {f.component}")

        # --- 3. Synthesise + write ---
        print("\n[3/3] build_report (GPT synthesis + xlsx write)…")
        report_path = build_report(
            findings=findings,
            images=image_assets,
            transcript=transcript,
            template_path=template_path,
            output_path=output_path,
        )
        size_kb = report_path.stat().st_size / 1024
        print(f"      wrote {report_path} ({size_kb:.1f} KB)")

        # --- Verify by reading the file back ---
        print("\n--- Verifying output ---")
        wb = load_workbook(report_path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            print(f"\nSheet: {sheet.title}")
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                non_empty = [str(c) if c is not None else "" for c in row]
                if any(non_empty):
                    preview = " | ".join(c[:60] for c in non_empty)
                    print(f"  {row_idx}: {preview}")
                if row_idx >= 12:
                    print("  ... (truncated)")
                    break
        wb.close()

        # Copy to a stable location so the user can inspect it.
        target = Path("var/smoke_report.xlsx").resolve()
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(report_path.read_bytes())
        print(f"\nSaved a copy at: {target}")

    print("\nAll good.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
