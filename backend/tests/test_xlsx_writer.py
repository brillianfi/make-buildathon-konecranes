"""Unit tests for xlsx workbook writer — no Azure, no network."""

from pathlib import Path

from app.domain.report import ReportSheet, ReportWorkbook
from app.services.xlsx_writer import write_workbook
from openpyxl import load_workbook
from PIL import Image


def _workbook(*sheets: ReportSheet) -> ReportWorkbook:
    return ReportWorkbook(sheets=list(sheets))


def _sheet(title: str, *rows: list) -> ReportSheet:
    return ReportSheet(title=title, rows=list(rows))


def _write(report: ReportWorkbook, target: Path, **kw: object) -> Path:
    return write_workbook(report, target, **kw)  # type: ignore[arg-type]


class TestWriteWorkbook:
    def test_creates_xlsx_file(self, tmp_path: Path) -> None:
        out = tmp_path / "report.xlsx"
        _write(_workbook(_sheet("Summary")), out)
        assert out.exists()

    def test_single_sheet_with_data(self, tmp_path: Path) -> None:
        out = tmp_path / "report.xlsx"
        report = _workbook(_sheet("Findings", ["Component", "Severity"], ["Hook", "Low"]))
        _write(report, out)

        wb = load_workbook(out)
        ws = wb["Findings"]
        assert ws.cell(1, 1).value == "Component"
        assert ws.cell(1, 2).value == "Severity"
        assert ws.cell(2, 1).value == "Hook"
        assert ws.cell(2, 2).value == "Low"

    def test_multiple_sheets(self, tmp_path: Path) -> None:
        out = tmp_path / "report.xlsx"
        _write(
            _workbook(_sheet("Summary", ["Total", 3]), _sheet("Details", ["img.jpg"])),
            out,
        )
        wb = load_workbook(out)
        assert wb.sheetnames == ["Summary", "Details"]

    def test_empty_workbook_creates_report_sheet(self, tmp_path: Path) -> None:
        out = tmp_path / "report.xlsx"
        _write(ReportWorkbook(sheets=[]), out)
        wb = load_workbook(out)
        assert "Report" in wb.sheetnames

    def test_none_cells_written_as_empty(self, tmp_path: Path) -> None:
        out = tmp_path / "report.xlsx"
        report = _workbook(_sheet("S", ["A", None, "B"]))
        _write(report, out)
        wb = load_workbook(out)
        assert wb["S"].cell(1, 2).value is None

    def test_creates_parent_directories(self, tmp_path: Path) -> None:
        out = tmp_path / "sub" / "dir" / "report.xlsx"
        _write(_workbook(_sheet("S")), out)
        assert out.exists()

    def test_image_embedded_when_filename_matches(self, tmp_path: Path) -> None:
        img_path = tmp_path / "photo.jpg"
        Image.new("RGB", (10, 10), color="blue").save(img_path, "JPEG")

        out = tmp_path / "report.xlsx"
        report = _workbook(_sheet("S", ["photo.jpg"]))
        _write(report, out, images_by_filename={"photo.jpg": img_path})

        wb = load_workbook(out)
        ws = wb["S"]
        # Cell is cleared when an image is embedded.
        assert ws.cell(1, 1).value in (None, "")
        assert len(ws._images) == 1  # type: ignore[attr-defined]
