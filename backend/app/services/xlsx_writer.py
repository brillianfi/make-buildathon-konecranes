from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from app.domain.report import ReportWorkbook

_EMBED_WIDTH_PX = 240
_EMBED_HEIGHT_PX = 180
_EMBED_ROW_HEIGHT_PT = 140  # ~180px
_EMBED_COL_WIDTH = 36  # column width units (~250px)


def _normalise_to_jpeg(path: Path) -> BytesIO:
    """Re-encode a JPG/MPO/etc. as a clean JPEG so openpyxl can embed it.

    DJI photos are MPO under a `.JPG` extension; openpyxl chokes on the .mpo
    mimetype during workbook save. Going through Pillow strips the multi-image
    container and produces a single-frame JPEG.
    """
    with PILImage.open(path) as opened:
        rgb = opened.convert("RGB") if opened.mode != "RGB" else opened.copy()
    buf = BytesIO()
    rgb.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return buf


def write_workbook(
    report: ReportWorkbook,
    target: Path,
    images_by_filename: dict[str, Path] | None = None,
) -> Path:
    """Materialise the workbook. Cells whose value matches a known image
    filename are emptied and replaced by the embedded image at that anchor."""
    images_by_filename = images_by_filename or {}

    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    if not report.sheets:
        wb.create_sheet("Report")

    for sheet in report.sheets:
        ws = wb.create_sheet(title=sheet.title[:31] or "Sheet")
        for row_idx, row in enumerate(sheet.rows, start=1):
            row_inserts: list[tuple[int, Path]] = []
            output_row: list[Any] = []
            for col_idx, cell in enumerate(row, start=1):
                if isinstance(cell, str) and cell in images_by_filename:
                    row_inserts.append((col_idx, images_by_filename[cell]))
                    output_row.append("")
                else:
                    output_row.append(cell)
            ws.append(output_row)
            if row_inserts:
                ws.row_dimensions[row_idx].height = _EMBED_ROW_HEIGHT_PT
                for col_idx, path in row_inserts:
                    img = XLImage(_normalise_to_jpeg(path))
                    img.width = _EMBED_WIDTH_PX
                    img.height = _EMBED_HEIGHT_PX
                    anchor = f"{get_column_letter(col_idx)}{row_idx}"
                    ws.add_image(img, anchor)
                    ws.column_dimensions[get_column_letter(col_idx)].width = (
                        _EMBED_COL_WIDTH
                    )

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    return target
