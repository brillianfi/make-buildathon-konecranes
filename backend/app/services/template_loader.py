from pathlib import Path

from openpyxl import load_workbook

from app.core.config import get_settings
from app.core.logging import get_logger
from app.domain.report import ReportTemplateRef

log = get_logger(__name__)

_MAX_PREVIEW_ROWS = 30


def list_templates() -> list[ReportTemplateRef]:
    settings = get_settings()
    folder = Path(settings.templates_dir)
    if not folder.exists():
        log.warning("templates.dir_missing", path=str(folder))
        return []
    refs: list[ReportTemplateRef] = []
    for p in sorted(folder.glob("*.xlsx")):
        if p.name.startswith("~"):  # Excel lock files
            continue
        refs.append(ReportTemplateRef(filename=p.name, path=str(p), size_bytes=p.stat().st_size))
    return refs


def resolve_template(filename: str) -> Path:
    settings = get_settings()
    folder = Path(settings.templates_dir).resolve()
    candidate = (folder / filename).resolve()
    # Prevent path traversal.
    if folder not in candidate.parents and candidate != folder:
        raise ValueError(f"Template not under templates dir: {filename}")
    if not candidate.exists():
        raise FileNotFoundError(f"Template not found: {filename}")
    return candidate


def flatten_template_to_text(template_path: Path) -> str:
    """Render the workbook as plain text so the LLM can mirror its structure.

    For each sheet we emit the title and the first N rows. Empty rows are
    skipped to keep the prompt compact.
    """
    wb = load_workbook(template_path, read_only=True, data_only=True)
    parts: list[str] = []
    for sheet in wb.worksheets:
        parts.append(f"### Sheet: {sheet.title}")
        emitted = 0
        for row in sheet.iter_rows(values_only=True):
            if all(c is None or c == "" for c in row):
                continue
            cells = ["" if c is None else str(c) for c in row]
            parts.append("\t".join(cells))
            emitted += 1
            if emitted >= _MAX_PREVIEW_ROWS:
                parts.append("... (truncated)")
                break
        parts.append("")
    wb.close()
    return "\n".join(parts).strip()
